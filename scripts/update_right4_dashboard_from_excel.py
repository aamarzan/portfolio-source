from __future__ import annotations
import argparse
import json
import re
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

SITE_ORDER = ["CMCH", "RMCH", "DMCH", "SZMCH", "SOMCH", "PGIMER", "SGRDUHS", "GGSMCH"]
SITE_META = {
    "CMCH": {"label": "CMCH", "country": "Bangladesh"},
    "RMCH": {"label": "RMCH", "country": "Bangladesh"},
    "DMCH": {"label": "DMCH", "country": "Bangladesh"},
    "SZMCH": {"label": "SZMCH", "country": "Bangladesh"},
    "SOMCH": {"label": "SOMCH", "country": "Bangladesh"},
    "PGIMER": {"label": "PGIMER", "country": "India"},
    "SGRDUHS": {"label": "SGRDUHS", "country": "India"},
    "GGSMCH": {"label": "GGSMCH", "country": "India"},
}
TARGET_SCHEDULE = [
    ("2025-06-01", 0.0, 0.0),
    ("2025-06-16", 60.0, 18.2),
    ("2025-07-01", 120.0, 36.4),
    ("2025-07-16", 180.0, 54.6),
    ("2025-07-31", 240.0, 72.8),
    ("2025-08-15", 300.0, 91.0),
    ("2025-08-30", 360.0, 109.2),
    ("2025-09-15", 420.0, 127.4),
    ("2025-09-30", 480.0, 145.6),
    ("2025-10-15", 540.0, 163.8),
    ("2025-10-31", 610.0, 182.0),
    ("2025-11-15", 670.0, 200.2),
    ("2025-12-03", 740.0, 218.4),
    ("2025-12-16", 790.0, 236.6),
    ("2025-12-30", 850.0, 254.8),
    ("2026-01-14", 910.0, 273.0),
    ("2026-01-31", 975.0, 291.2),
    ("2026-02-15", 1040.0, 309.4),
    ("2026-03-10", 1135.0, 327.6),
    ("2026-03-28", 1200.0, 345.8),
    ("2026-04-12", 1260.0, 364.0),
    ("2026-04-27", 1320.0, 382.2),
    ("2026-05-12", 1380.0, 400.4),
    ("2026-05-27", 1440.0, 418.6),
    ("2026-06-11", 1500.0, 436.8),
    ("2026-06-26", 1560.0, 455.0),
    ("2026-07-11", 1620.0, 474.0),
]
TRUE_POSITIVE_TARGET_SCHEDULE = [
    ("2025-06-01", 0.0),
    ("2025-06-16", 3.0),
    ("2025-07-01", 6.0),
    ("2025-07-16", 9.0),
    ("2025-07-31", 12.0),
    ("2025-08-15", 15.0),
    ("2025-08-30", 18.0),
    ("2025-09-15", 21.0),
    ("2025-09-30", 24.0),
    ("2025-10-15", 27.0),
    ("2025-10-31", 30.0),
    ("2025-11-15", 33.0),
    ("2025-12-03", 36.5),
    ("2025-12-16", 39.0),
    ("2025-12-30", 42.0),
    ("2026-01-14", 45.0),
    ("2026-01-30", 48.0),
    ("2026-02-17", 51.5),
    ("2026-03-10", 55.7),
    ("2026-03-17", 57.0),
    ("2026-04-02", 60.0),
    ("2026-04-17", 63.0),
    ("2026-05-02", 66.0),
    ("2026-05-17", 69.0),
    ("2026-06-02", 72.0),
    ("2026-06-17", 75.0),
    ("2026-07-02", 78.0),
    ("2026-07-17", 81.0),
]


def normalize_site(code: str) -> str:
    code = (code or "").strip().upper()
    return {"SGRDUSH": "SGRDUHS", "GGSMC": "GGSMCH"}.get(code, code)


def load_master_dataframe(workbook_path: Path) -> pd.DataFrame:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    df = pd.DataFrame(rows, columns=headers)
    df = df[df["Patient ID"].notna()].copy()

    df["Date of Screening"] = pd.to_datetime(df["Date of Screening"], errors="coerce")
    df = df[df["Date of Screening"].notna()].copy()

    df["siteCode"] = (
        df["Patient ID"].astype(str).str.extract(r"^([A-Za-z]+)")[0].str.upper().map(normalize_site)
    )
    df["siteCode"] = df["siteCode"].fillna("")

    for col in ["Patient Status", "Outcome (Died/ Survived)", "If excluded, reason", "True Positive?", "Comments"]:
        if col in df.columns:
            df[col] = df[col].fillna("").astype(str).str.strip()

    df["isEnrolled"] = df["Patient Status"].str.lower().eq("enrolled")
    df["isExcluded"] = df["Patient Status"].str.lower().eq("excluded")
    df["isDiedEnrolled"] = df["Outcome (Died/ Survived)"].str.lower().eq("died") & df["isEnrolled"]
    df["isTruePositive"] = df["True Positive?"].str.lower().eq("yes")
    return df


def safe_base_deficit(value) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value)


def build_records(df: pd.DataFrame) -> list[dict]:
    records: list[dict] = []
    for _, row in df.sort_values(["Date of Screening", "Patient ID"], ascending=[True, True]).iterrows():
        site_code = row["siteCode"]
        meta = SITE_META.get(site_code, {"label": site_code or "Unknown", "country": "Other"})
        records.append({
            "patientId": str(row["Patient ID"]),
            "siteCode": site_code,
            "siteLabel": meta["label"],
            "country": meta["country"],
            "screeningDate": row["Date of Screening"].date().isoformat(),
            "baseDeficit": safe_base_deficit(row.get("Base Deficit Value")),
            "patientStatus": row.get("Patient Status", ""),
            "outcome": row.get("Outcome (Died/ Survived)", ""),
            "truePositive": "Yes" if bool(row["isTruePositive"]) else "No",
            "comment": row.get("Comments", "") or "",
            "excludedReason": row.get("If excluded, reason", "") or "",
            "isEnrolled": bool(row["isEnrolled"]),
            "isExcluded": bool(row["isExcluded"]),
            "isDiedEnrolled": bool(row["isDiedEnrolled"]),
            "isTruePositive": bool(row["isTruePositive"]),
        })
    return records


def build_payload(df: pd.DataFrame) -> dict:
    total_recruited = int(df["isEnrolled"].sum())
    total_true_positive = int(df["isTruePositive"].sum())
    total_screened = int(len(df))
    total_excluded = int(df["isExcluded"].sum())
    total_died_enrolled = int(df["isDiedEnrolled"].sum())
    latest_date = df["Date of Screening"].max().date().isoformat() if not df.empty else ""
    latest_enrolled_date = df.loc[df["isEnrolled"], "Date of Screening"].max().date().isoformat() if df["isEnrolled"].any() else ""
    latest_true_positive_date = df.loc[df["isTruePositive"], "Date of Screening"].max().date().isoformat() if df["isTruePositive"].any() else ""

    payload = {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceWorkbook": "private/master_data.xlsx",
            "pageTitle": "NIHR RIGHT4 Methanol Dashboard",
            "pageSubtitle": "Operational recruitment, screening, and true-positive monitoring synced from the master workbook.",
            "latestScreeningDate": latest_date,
            "latestEnrolledDate": latest_enrolled_date,
            "latestTruePositiveDate": latest_true_positive_date,
        },
        "config": {
            "siteOrder": SITE_ORDER,
            "siteMeta": SITE_META,
            "targetSchedule": {
                "dates": [d for d, _, _ in TARGET_SCHEDULE],
                "targetPatients": [t for _, t, _ in TARGET_SCHEDULE],
                "calculatedTarget": [c for _, _, c in TARGET_SCHEDULE],
            },
            "truePositiveTargetSchedule": {
                "dates": [d for d, _ in TRUE_POSITIVE_TARGET_SCHEDULE],
                "targetPositive": [t for _, t in TRUE_POSITIVE_TARGET_SCHEDULE],
            },
            "studyTargets": {
                "overallRecruitmentTarget": 1620,
                "truePositiveTarget": 81,
            },
        },
        "summary": {
            "totalScreened": total_screened,
            "totalRecruited": total_recruited,
            "totalExcluded": total_excluded,
            "totalDiedEnrolled": total_died_enrolled,
            "totalTruePositive": total_true_positive,
            "overallPositivePct": (total_true_positive / total_recruited * 100) if total_recruited else 0,
            "latestEnrolledDate": latest_enrolled_date,
            "latestTruePositiveDate": latest_true_positive_date,
        },
        "records": build_records(df),
    }
    return payload


def update_index_versions(index_path: Path, version: str) -> None:
    if not index_path.exists():
        return
    html = index_path.read_text(encoding="utf-8")
    patterns = [
        r"(right4-dashboard\.css\?v=)([^\"']+)",
        r"(right4-dashboard-data\.json\?v=)([^\"']+)",
        r"(right4-dashboard-data\.js\?v=)([^\"']+)",
        r"(right4-dashboard\.js\?v=)([^\"']+)",
    ]
    for pattern in patterns:
        html = re.sub(pattern, rf"\g<1>{version}", html)
    index_path.write_text(html, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--version", default=datetime.now().strftime("%Y%m%d-%H%M%S"))
    args = parser.parse_args()

    repo_root = Path(__file__).resolve().parents[1]
    workbook_path = repo_root / "private" / "master_data.xlsx"
    index_path = repo_root / "right4-dashboard" / "index.html"
    data_js_path = repo_root / "right4-dashboard-data.js"
    data_json_path = repo_root / "right4-dashboard-data.json"

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    df = load_master_dataframe(workbook_path)
    payload = build_payload(df)

    data_js_path.write_text(
        "window.RIGHT4_DASHBOARD_DATA = " + json.dumps(payload, indent=2, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    data_json_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    update_index_versions(index_path, args.version)

    print(f"Workbook rows: {len(df)}")
    print(f"Total screened: {payload['summary']['totalScreened']}")
    print(f"Total recruited: {payload['summary']['totalRecruited']}")
    print(f"Total true positive: {payload['summary']['totalTruePositive']}")
    print(f"Last updated: {payload['meta']['generatedAt']}")
    print(f"Latest screening date in workbook: {payload['meta']['latestScreeningDate']}")
    print(f"Latest enrolled date in workbook: {payload['meta']['latestEnrolledDate'] or '-'}")
    print(f"Latest true positive date in workbook: {payload['meta']['latestTruePositiveDate'] or '-'}")


if __name__ == "__main__":
    main()
