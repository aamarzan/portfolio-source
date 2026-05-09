from __future__ import annotations

import argparse
import csv
import json
import math
import re
from datetime import datetime, date, time
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

SITE_ORDER = ["CMCH", "RMCH", "DMCH", "SZMCH", "SOMCH", "CoxMCH", "PGIMER", "SGRDUHS", "GGSMCH"]
SITE_META = {
    "CMCH": {"label": "CMCH", "country": "Bangladesh"},
    "RMCH": {"label": "RMCH", "country": "Bangladesh"},
    "DMCH": {"label": "DMCH", "country": "Bangladesh"},
    "SZMCH": {"label": "SZMCH", "country": "Bangladesh"},
    "SOMCH": {"label": "SOMCH", "country": "Bangladesh"},
    "CoxMCH": {"label": "CoxMCH", "country": "Bangladesh"},
    "PGIMER": {"label": "PGIMER", "country": "India"},
    "SGRDUHS": {"label": "SGRDUHS", "country": "India"},
    "GGSMCH": {"label": "GGSMCH", "country": "India"},
}
CLASS_ORDER = ["True positive", "False positive", "False negative", "True negative", "Unclassified"]
SEV_ORDER = [
    "Negative",
    "Low positive",
    "Moderate positive",
    "High positive",
    "Positive, category not specified",
    "Indeterminate",
    "Missing",
]
INTERPRETATION = [
    {"category": "Negative", "formateRange": "0–1.99 mmol/L", "clinicalMeaning": "Normal formate level. Re-test if ethanol co-ingestion may delay methanol metabolism."},
    {"category": "Low positive", "formateRange": "2–4.99 mmol/L", "clinicalMeaning": "Slightly elevated formate level, typically early-stage or smaller exposure. Re-test after 2 hours."},
    {"category": "Moderate positive", "formateRange": "5–10 mmol/L", "clinicalMeaning": "Moderately elevated formate level. Patients may be symptomatic or soon become symptomatic."},
    {"category": "High positive", "formateRange": ">10 mmol/L", "clinicalMeaning": "Strongly elevated formate level. Immediate specific treatment is usually required."},
]


def txt(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.strftime("%H:%M")
    return str(value).strip()


def normalized_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", txt(value).lower())


def numeric_value(value: Any) -> float | None:
    raw = txt(value).replace(",", "")
    if not raw:
        return None
    m = re.search(r"[-+]?\d+(?:\.\d+)?", raw)
    if not m:
        return None
    try:
        return float(m.group(0))
    except ValueError:
        return None


def date_iso(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = txt(value)
    if not raw:
        return ""
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", raw):
        return raw
    for fmt in ("%A, %B %d, %Y", "%B %d, %Y", "%d %B %Y", "%d-%b-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    # final permissive attempt without adding a dependency
    try:
        from dateutil import parser as date_parser  # type: ignore
        return date_parser.parse(raw, dayfirst=False).date().isoformat()
    except Exception:
        return raw


def yes(value: Any) -> bool:
    return normalized_key(value) in {"yes", "y", "true", "tp", "truepositive", "positive"}


def find_column(headers: list[str], kind: str) -> str | None:
    exact = {normalized_key(h): h for h in headers if txt(h)}
    candidates = {
        "tp": ["truepositive", "truepositive?", "istruepositive", "tp"],
        "poc": [
            "pocresult", "poctestresult", "poc", "pocformate", "pocformateresult",
            "pointofcareresult", "pointofcaretestresult", "pointofcareformate", "pocassayresult",
        ],
        "lab": [
            "labresult", "laboratoryresult", "labassayresult", "labformate", "labformateresult",
            "formallabassay", "formallabresult", "formateassay", "formateresult", "referenceresult",
            "referenceassay", "confirmatorylabresult",
        ],
    }[kind]
    for candidate in candidates:
        if candidate in exact:
            return exact[candidate]
    for h in headers:
        k = normalized_key(h)
        if kind == "tp" and "true" in k and "positive" in k:
            return h
        if kind == "poc" and ("poc" in k or "pointofcare" in k) and any(x in k for x in ["result", "assay", "test", "formate", "call"]):
            return h
        if kind == "lab" and any(x in k for x in ["lab", "laboratory", "formal", "reference", "confirmatory", "assay"]) and any(x in k for x in ["result", "formate", "value", "call"]) and "poc" not in k:
            return h
    return None


def detect_columns(headers: list[str]) -> tuple[str | None, str | None, str | None, str]:
    tp_col = find_column(headers, "tp")
    poc_col = find_column(headers, "poc")
    lab_col = find_column(headers, "lab")
    source = "header-name detection"
    non_empty = [h for h in headers if txt(h)]
    # RIGHT4 convention: ... True Positive?, POC Result, Lab Result at the rightmost end.
    if (not tp_col or not poc_col or not lab_col) and len(non_empty) >= 12:
        right3 = non_empty[-3:]
        if normalized_key(right3[0]) in {"truepositive", "truepositive?"} or ("true" in normalized_key(right3[0]) and "positive" in normalized_key(right3[0])):
            tp_col = tp_col or right3[0]
            poc_col = poc_col or right3[1]
            lab_col = lab_col or right3[2]
            source = "rightmost-three fallback"
    return tp_col, poc_col, lab_col, source


def normalize_site(patient_id: Any) -> str:
    m = re.match(r"^([A-Za-z]+)", txt(patient_id))
    raw = m.group(1) if m else ""
    upper = raw.upper()
    return {
        "SGRDUSH": "SGRDUHS",
        "GGSMC": "GGSMCH",
        "GGSMCH": "GGSMCH",
        "COXMCH": "CoxMCH",
        "COXMHC": "CoxMCH",
    }.get(upper, raw if raw == "CoxMCH" else upper)


def load_excel_rows(workbook_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [txt(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
    rows: list[dict[str, Any]] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if txt(row.get("Patient ID")):
            rows.append(row)
    return headers, rows


def load_tsv_rows(tsv_path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    if not tsv_path.exists():
        return [], []
    text = tsv_path.read_text(encoding="utf-8-sig")
    lines = [line.rstrip("\r\n") for line in text.splitlines() if line.strip()]
    if not lines:
        return [], []
    reader = csv.DictReader(lines, delimiter="\t")
    headers = reader.fieldnames or []
    rows = [row for row in reader if txt(row.get("Patient ID"))]
    return headers, rows


def source_rows(repo_root: Path) -> tuple[list[str], list[dict[str, Any]], str, str | None, str | None, str | None]:
    workbook_path = repo_root / "private" / "master_data.xlsx"
    headers, rows = load_excel_rows(workbook_path)
    tp_col, poc_col, lab_col, detection_source = detect_columns(headers)
    if poc_col and lab_col:
        return headers, rows, detection_source, tp_col, poc_col, lab_col

    # If the workbook uploaded to this environment does not contain the rightmost POC/Lab fields,
    # use the pasted TSV source as a patient-level fallback. This keeps the POC page functional
    # while preserving workbook-first behaviour whenever the real workbook has the columns.
    fallback_path = repo_root / "private" / "poc_results_source.tsv"
    fallback_headers, fallback_rows = load_tsv_rows(fallback_path)
    f_tp, f_poc, f_lab, f_source = detect_columns(fallback_headers)
    if f_poc and f_lab and fallback_rows:
        return fallback_headers, fallback_rows, f"fallback TSV source ({f_source})", f_tp, f_poc, f_lab

    return headers, rows, detection_source, tp_col, poc_col, lab_col


def poc_severity(value: Any) -> str:
    k = normalized_key(value)
    n = numeric_value(value)
    if not k or k in {"na", "n/a", "notapplicable", "notavailable", "notconducted", "notdone", "pending", "missing"}:
        return "Missing"
    if ("negative" in k or k in {"normal", "neg", "notdetected", "undetected", "nonreactive"}) and "positive" not in k and "pos" not in k:
        return "Negative"
    if any(x in k for x in ["highpositive", "highpos", "strongpositive", "strongpos"]):
        return "High positive"
    if any(x in k for x in ["moderatepositive", "moderatepos", "mediumpositive", "mediumpos", "medpositive", "medpos", "modpositive", "modpos"]):
        return "Moderate positive"
    if any(x in k for x in ["lowpositive", "lowpos", "slightlypositive", "weakpositive", "weakpos"]):
        return "Low positive"
    if n is not None:
        if n < 2.0:
            return "Negative"
        if n < 5.0:
            return "Low positive"
        if n <= 10.0:
            return "Moderate positive"
        return "High positive"
    if any(x in k for x in ["positive", "pos", "yes", "detected", "reactive"]):
        return "Positive, category not specified"
    return "Indeterminate"


def lab_severity(value: Any) -> str:
    k = normalized_key(value)
    n = numeric_value(value)
    if not k or k in {"na", "n/a", "notapplicable", "notavailable", "notconducted", "notdone", "pending", "missing"}:
        return "Missing"
    if n is not None:
        if n < 2.0:
            return "Negative"
        if n < 5.0:
            return "Low positive"
        if n <= 10.0:
            return "Moderate positive"
        return "High positive"
    return poc_severity(value)


def diagnostic_call(severity: str) -> str:
    if severity == "Missing":
        return "Missing"
    if severity == "Negative":
        return "Negative"
    if severity in {"Low positive", "Moderate positive", "High positive", "Positive, category not specified"}:
        return "Positive"
    return "Indeterminate"


def classify_pair(poc_result: Any, lab_result: Any) -> tuple[str, str, str, str, str, str, bool]:
    ps = poc_severity(poc_result)
    ls = lab_severity(lab_result)
    pc = diagnostic_call(ps)
    lc = diagnostic_call(ls)
    if pc == "Positive" and lc == "Positive":
        return "True positive", pc, lc, ps, ls, "Lab ≥2.00 mmol/L and POC positive.", True
    if pc == "Negative" and lc == "Positive":
        return "False negative", pc, lc, ps, ls, "Lab ≥2.00 mmol/L but POC negative.", True
    if pc == "Positive" and lc == "Negative":
        return "False positive", pc, lc, ps, ls, "Lab <2.00 mmol/L but POC positive.", True
    if pc == "Negative" and lc == "Negative":
        return "True negative", pc, lc, ps, ls, "Lab <2.00 mmol/L and POC negative.", True
    return "Unclassified", pc, lc, ps, ls, "Missing or uninterpretable POC/lab pair.", False


def calculate_metrics(counts: dict[str, int]) -> dict[str, float | None]:
    tp = counts.get("True positive", 0)
    fp = counts.get("False positive", 0)
    fn = counts.get("False negative", 0)
    tn = counts.get("True negative", 0)
    def safe(num: float, den: float) -> float | None:
        return num / den if den else None
    sens = safe(tp, tp + fn)
    spec = safe(tn, tn + fp)
    ppv = safe(tp, tp + fp)
    npv = safe(tn, tn + fn)
    acc = safe(tp + tn, tp + fp + fn + tn)
    f1 = safe(2 * tp, 2 * tp + fp + fn)
    denom = math.sqrt((tp + fp) * (tp + fn) * (tn + fp) * (tn + fn))
    mcc = ((tp * tn) - (fp * fn)) / denom if denom else None
    youden = (sens + spec - 1) if sens is not None and spec is not None else None
    return {"sensitivity": sens, "specificity": spec, "ppv": ppv, "npv": npv, "accuracy": acc, "f1": f1, "mcc": mcc, "youden": youden}


def build_payload(repo_root: Path) -> dict[str, Any]:
    headers, rows, detection_source, tp_col, poc_col, lab_col = source_rows(repo_root)
    records: list[dict[str, Any]] = []
    class_counts = {cls: 0 for cls in CLASS_ORDER}
    true_positive_declarations = 0

    for row in rows:
        patient_id = txt(row.get("Patient ID"))
        if not patient_id:
            continue
        site_code = normalize_site(patient_id)
        meta = SITE_META.get(site_code, {"label": site_code or "Unknown", "country": "Other"})
        true_positive_raw = txt(row.get(tp_col)) if tp_col else ""
        poc_raw = txt(row.get(poc_col)) if poc_col else ""
        lab_raw = txt(row.get(lab_col)) if lab_col else ""
        if yes(true_positive_raw):
            true_positive_declarations += 1

        diagnostic_class, poc_call, lab_call, poc_sev, lab_sev, note, is_classifiable = classify_pair(poc_raw, lab_raw)
        class_counts[diagnostic_class] = class_counts.get(diagnostic_class, 0) + 1
        records.append({
            "patientId": patient_id,
            "siteCode": site_code,
            "siteLabel": meta["label"],
            "country": meta["country"],
            "screeningDate": date_iso(row.get("Date of Screening")),
            "screeningDateRaw": txt(row.get("Date of Screening")),
            "patientStatus": txt(row.get("Patient Status")),
            "outcome": txt(row.get("Outcome (Died/ Survived)")),
            "baseDeficit": txt(row.get("Base Deficit Value")),
            "truePositiveRaw": true_positive_raw,
            "isMasterTruePositive": yes(true_positive_raw),
            "pocResult": poc_raw,
            "labResult": lab_raw,
            "labNumeric": numeric_value(lab_raw),
            "pocCall": poc_call,
            "labCall": lab_call,
            "pocSeverity": poc_sev,
            "labSeverity": lab_sev,
            "diagnosticClass": diagnostic_class,
            "isClassifiable": is_classifiable,
            "classificationNote": note,
            "comments": txt(row.get("Comments")),
        })

    records.sort(key=lambda r: (r.get("screeningDate") or "", r.get("patientId") or ""))
    classifiable_rows = sum(v for k, v in class_counts.items() if k != "Unclassified")
    metrics = calculate_metrics(class_counts)
    latest_screening = max([r["screeningDate"] for r in records if r.get("screeningDate")], default="")

    return {
        "meta": {
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
            "sourceWorkbook": "private/master_data.xlsx",
            "pageTitle": "POC Performance Analytics",
            "pageSubtitle": "Diagnostic performance of point-of-care formate testing against formal laboratory formate assay.",
            "latestScreeningDate": latest_screening,
            "workbookColumns": headers,
            "columnDetectionSource": detection_source,
            "pocColumnDetected": poc_col or "",
            "labColumnDetected": lab_col or "",
            "truePositiveColumnDetected": tp_col or "",
            "hasPocColumn": bool(poc_col),
            "hasLabColumn": bool(lab_col),
            "hasUsablePocData": classifiable_rows > 0,
        },
        "config": {
            "siteOrder": SITE_ORDER,
            "siteMeta": SITE_META,
            "severityOrder": SEV_ORDER,
            "classOrder": CLASS_ORDER,
            "interpretation": INTERPRETATION,
            "thresholdRule": "Lab ≥2.00 mmol/L is positive; lab <2.00 mmol/L is negative. POC low, moderate/medium, or high positive is positive.",
        },
        "summary": {
            "totalRows": len(records),
            "classifiableRows": classifiable_rows,
            "unclassifiedRows": class_counts.get("Unclassified", 0),
            "classCounts": class_counts,
            "truePositiveDeclarations": true_positive_declarations,
            "metrics": metrics,
        },
        "records": records,
    }


def update_versions(index_path: Path, version: str) -> None:
    s = index_path.read_text(encoding="utf-8")
    for pat in [
        r"(right4-poc-performance\.css\?v=)([^\"']+)",
        r"(right4-poc-performance-data\.json\?v=)([^\"']+)",
        r"(right4-poc-performance-data\.js\?v=)([^\"']+)",
        r"(right4-poc-performance\.js\?v=)([^\"']+)",
        r"(right4-access\.js\?v=)([^\"']+)",
    ]:
        s = re.sub(pat, rf"\g<1>{version}", s)
    index_path.write_text(s, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--version", default=datetime.now().strftime("%Y%m%d-%H%M%S"))
    args = parser.parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    payload = build_payload(repo_root)
    (repo_root / "right4-poc-performance-data.js").write_text(
        "window.RIGHT4_POC_PERFORMANCE_DATA = " + json.dumps(payload, indent=2, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    (repo_root / "right4-poc-performance-data.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    update_versions(repo_root / "right4-poc-performance" / "index.html", args.version)

    counts = payload["summary"]["classCounts"]
    metrics = payload["summary"]["metrics"]
    print(f"Rows: {payload['summary']['totalRows']}")
    print(f"Classifiable: {payload['summary']['classifiableRows']}")
    print(f"Unclassified: {payload['summary']['unclassifiedRows']}")
    print(f"True Positive? declarations: {payload['summary']['truePositiveDeclarations']}")
    print(f"TP={counts['True positive']} FP={counts['False positive']} FN={counts['False negative']} TN={counts['True negative']}")
    print(f"Sensitivity={metrics['sensitivity']:.4f}" if metrics['sensitivity'] is not None else "Sensitivity=N/A")
    print(f"Specificity={metrics['specificity']:.4f}" if metrics['specificity'] is not None else "Specificity=N/A")
    print(f"POC column: {payload['meta']['pocColumnDetected'] or '-'}")
    print(f"Lab column: {payload['meta']['labColumnDetected'] or '-'}")
    print(f"Detection source: {payload['meta']['columnDetectionSource']}")


if __name__ == "__main__":
    main()
