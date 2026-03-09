#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
from collections import Counter
from pathlib import Path

import openpyxl

SITE_ALIASES = {
    "SGRDUSH": "SGRDUHS",
}

SITE_NAMES = {
    "CMCH": "Chattogram Medical College Hospital, CMCH",
    "DMCH": "Dhaka Medical College Hospital, DMCH",
    "PGIMER": "Postgraduate Institute of Medical Education and Research, PGIMER",
    "RMCH": "Rajshahi Medical College Hospital, RMCH",
    "SOMCH": "Sylhet MAG Osmani Medical College Hospital, SOMCH",
    "SZMCH": "Shaheed Ziaur Rahman Medical College Hospital, SZMCH",
    "GGSMCH": "Guru Gobind Singh Medical College Hospital, GGSMCH",
    "SGRDUHS": "SRI GURU RAM DAS UNIVERSITY OF HEALTH SCIENCES, SGRDUHS",
}

SITE_ORDER = ["PGIMER", "RMCH", "DMCH", "CMCH", "SZMCH", "GGSMCH", "SGRDUHS", "SOMCH"]
THEME_ORDER = [
    "Case classification & source logic",
    "Outcome & follow-up consistency",
    "Consent & eligibility",
    "Chronology & timing",
    "Clinical values & investigations",
    "Treatment regimen / administration",
    "Duplicate / repeated records",
    "Laboratory / image evidence",
    "Exposure details & volume",
    "Demographics & field completeness",
    "Other / operational review",
]
STATUS_ORDER = ["Open", "Not entered", "Resolved"]
SEVERITY_ORDER = ["Major", "Moderate", "Minor", "Standard"]


def clean_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_site_code(patient_id: str) -> str:
    code = clean_text(patient_id).split("-")[0].upper()
    return SITE_ALIASES.get(code, code)


def to_iso_date(value):
    if value in (None, ""):
        return None
    if hasattr(value, "date") and hasattr(value, "year") and hasattr(value, "month"):
        try:
            return value.date().isoformat()
        except Exception:
            try:
                return value.isoformat()
            except Exception:
                pass
    text = clean_text(value)
    if not text:
        return None

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m/%d/%y", "%d-%m-%Y", "%m-%d-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except Exception:
            continue

    return text


def normalize_severity(value) -> str:
    s = clean_text(value).lower()
    if s.startswith("maj"):
        return "Major"
    if s.startswith("mod"):
        return "Moderate"
    if s.startswith("min"):
        return "Minor"
    return "Standard"


def normalize_status(value, data_entry_date=None) -> str:
    s = clean_text(value).lower()
    if s in {"yes", "resolved", "done", "closed"}:
        return "Resolved"
    if s in {"pending", "open", "no"}:
        return "Open"
    if not s and not data_entry_date:
        return "Not entered"
    return "Open"


def detect_explicit_screen(query: str):
    q = clean_text(query)
    m = re.search(r"\b(Recruit(?:e?m?e?n?t|uitemnet)?|Assessment)\s+Screen\s+(\d+)(\s*\([^)]*\))?", q, re.I)
    if m:
        form_raw = m.group(1).lower()
        form = "Assessment" if "assessment" in form_raw else "Recruitment"
        screen = f"{form} Screen {m.group(2)}{m.group(3) or ''}"
        return form, screen.strip()

    m = re.search(r"\bAll\s+Assessments?\s*:\s*screen\s*(\d+)", q, re.I)
    if m:
        return "Assessment", f"Assessment Screen {m.group(1)}"

    return None, None


def infer_form_and_screen(query: str):
    q = clean_text(query)
    ql = q.lower()
    explicit_form, explicit_screen = detect_explicit_screen(q)
    if explicit_form:
        return explicit_form, explicit_screen

    if "first assessment" in ql:
        return "Assessment", "First assessment"
    if "outcome assessment" in ql:
        return "Assessment", "Outcome assessment"

    num_match = re.search(r"\bscreen\s*(\d+)\b", ql)
    num = num_match.group(1) if num_match else None

    assessment_kw = [
        "assessment", "bicarbonate", "fomepizole", "folic acid", "loading dose",
        "bolus", "iv fluid", "last review", "review type", "review time", "review date",
        "was bicarbonate", "dose of", "participant outcome", "sequela", "cascade assessments",
        "treatment level", "outcome time", "outcome date", "discharged", "absconded", "inpatient"
    ]
    recruitment_kw = [
        "recruitment", "enrollment", "enrolled", "consent", "poisoning type", "suspected methanol",
        "metabolic acidosis", "poison volume", "known source", "source is liquor", "poison class",
        "intake of illegal", "bootleg", "spurious alcohol", "start of symptoms", "onset",
        "time of consumption", "investigation", "poc", "lab date", "lab details", "lab result",
        "family name", "division", "district", "reg number", "blood pressure", "base excess",
        "ph ", "hco3", "potassium"
    ]

    assess_count = sum(1 for k in assessment_kw if k in ql)
    recruit_count = sum(1 for k in recruitment_kw if k in ql)

    if assess_count > recruit_count:
        form = "Assessment"
    elif recruit_count > assess_count:
        form = "Recruitment"
    elif assess_count and recruit_count:
        form = "Cross-form"
    else:
        form = "Assessment" if "assessment" in ql else "Recruitment" if "recruit" in ql else "General"

    if num:
        if form in {"Recruitment", "Assessment"}:
            return form, f"{form} Screen {num}"
        if assess_count > 0 and recruit_count == 0:
            return "Assessment", f"Assessment Screen {num}"
        if recruit_count > 0 and assess_count == 0:
            return "Recruitment", f"Recruitment Screen {num}"
        return form, "General"

    if form == "Assessment" and "assessment" in ql:
        return form, "General"
    return form, "General"


def infer_theme(query: str, form: str, screen: str) -> str:
    q = clean_text(query)
    ql = q.lower()

    if any(x in ql for x in ["duplicate", "duplicated", "2 records", "two assessment records"]):
        return "Duplicate / repeated records"

    if any(x in ql for x in ["consent", "lacking capacity", "inclusion criteria", "eligible"]):
        return "Consent & eligibility"

    if any(x in ql for x in ["outcome", "discharge", "discharged", "absconded", "remained admitted", "inpatient", "sequela", "follow-up"]):
        return "Outcome & follow-up consistency"

    if any(x in ql for x in ["fomepizole", "folic acid", "bicarbonate", "loading dose", "bolus", "iv fluid", "administered", "treatment", "stopped", "restarted"]):
        return "Treatment regimen / administration"

    if any(x in ql for x in ["image", "images", "lab result in the image", "lab details if already attained", "earlier lab result image"]):
        return "Laboratory / image evidence"

    if "poison volume" in ql:
        return "Exposure details & volume"

    if any(x in ql for x in ["division", "district", "family name", "reg number", "fields should not be empty", "fill up as much information", "not-recruited cases"]):
        return "Demographics & field completeness"

    if any(x in ql for x in ["suspected methanol", "metabolic acidosis", "poisoning type", "source is liquor", "known source", "poison class", "illegal/bootleg/spurious alcohol", "contradictory"]):
        return "Case classification & source logic"

    if any(x in ql for x in ["ph ", "hco3", "base excess", "base deficit", "potassium", "blood pressure", "systolic", "diastolic"]):
        return "Clinical values & investigations"

    if any(x in ql for x in ["date", "time", "chronology", "before", "after", "onset", "start of symptoms", "review time", "review date", "investigation date", "investigation time", "poc date", "lab date", "intake to symptoms", "duration"]):
        return "Chronology & timing"

    if any(x in ql for x in ["investigation", "poc", "lab"]):
        return "Clinical values & investigations" if ("screen 7" in ql or "screen 9" in ql) else "Chronology & timing"

    return "Other / operational review"


def find_columns(ws):
    headers = [clean_text(ws.cell(1, c).value).lower() for c in range(1, ws.max_column + 1)]
    header_map = {name: idx + 1 for idx, name in enumerate(headers)}

    def col(*fragments):
        for fragment in fragments:
            for key, idx in header_map.items():
                if fragment in key:
                    return idx
        raise KeyError(f"Could not find a column matching: {fragments}")

    return {
        "patient_id": col("patient id"),
        "severity": col("severity"),
        "query": col("query"),
        "data_entry_date": col("data entry"),
        "query_sent_date": col("query sent"),
        "status": col("resolve status", "status"),
    }


def build_payload(workbook_path: Path):
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cols = find_columns(ws)

    rows = []
    participant_ids = set()
    site_counts = Counter()
    site_codes_seen = []

    for r in range(2, ws.max_row + 1):
        patient_id = ws.cell(r, cols["patient_id"]).value
        query = ws.cell(r, cols["query"]).value
        if not patient_id or not query:
            continue

        site_code = normalize_site_code(patient_id)
        if site_code not in site_counts:
            site_codes_seen.append(site_code)
        site_counts[site_code] += 1
        participant_ids.add(clean_text(patient_id).upper())

        form, screen = infer_form_and_screen(query)
        theme = infer_theme(query, form, screen)
        severity = normalize_severity(ws.cell(r, cols["severity"]).value)
        data_entry = to_iso_date(ws.cell(r, cols["data_entry_date"]).value)
        query_sent = to_iso_date(ws.cell(r, cols["query_sent_date"]).value)
        status_group = normalize_status(ws.cell(r, cols["status"]).value, data_entry)

        rows.append({
            "siteCode": site_code,
            "siteName": SITE_NAMES.get(site_code, site_code),
            "query": clean_text(query),
            "form": form,
            "severity": severity,
            "theme": theme,
            "screen": screen or "General",
            "status_group": status_group,
            "querySentDate": query_sent,
            "dataEntryDate": data_entry,
            "count": 1,
        })

    ordered_sites = [s for s in SITE_ORDER if s in site_counts] + [s for s in site_codes_seen if s not in SITE_ORDER]
    dated_batches = [
        row["querySentDate"]
        for row in rows
        if row["querySentDate"] and re.match(r"^\d{4}-\d{2}-\d{2}$", str(row["querySentDate"]))
    ]
    last_query_sent = max(dated_batches) if dated_batches else None

    meta = {
        "title": "Methanol Trial eCRF Query Operations Dashboard",
        "totalQueries": len(rows),
        "activeSites": len(site_counts),
        "totalParticipants": len(participant_ids),
        "lastQuerySent": last_query_sent,
        "openQueries": sum(1 for row in rows if row["status_group"] == "Open"),
        "resolvedQueries": sum(1 for row in rows if row["status_group"] == "Resolved"),
        "notEnteredQueries": sum(1 for row in rows if row["status_group"] == "Not entered"),
        "siteCounts": {code: site_counts[code] for code in ordered_sites},
        "siteNames": {code: SITE_NAMES.get(code, code) for code in ordered_sites},
    }

    return {
        "META": meta,
        "DATA": rows,
        "SITE_ORDER": ordered_sites,
        "THEME_ORDER": THEME_ORDER,
        "STATUS_ORDER": STATUS_ORDER,
        "SEVERITY_ORDER": SEVERITY_ORDER,
        "GENERATED_AT": dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "SOURCE_WORKBOOK": workbook_path.name,
    }


def write_outputs(payload: dict, output_js: Path, output_json: Path):
    output_json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    output_js.write_text("window.ECRF_PAYLOAD = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")


def bump_asset_versions(index_path: Path, version_token: str):
    if not index_path.exists():
        return

    text = index_path.read_text(encoding="utf-8")

    replacements = {
        r"(ecrf\.css\?v=)[^\"']+": rf"\g<1>{version_token}",
        r"(ecrf-data\.js\?v=)[^\"']+": rf"\g<1>{version_token}",
        r"(ecrf\.js\?v=)[^\"']+": rf"\g<1>{version_token}",
    }

    for pattern, replacement in replacements.items():
        text = re.sub(pattern, replacement, text)

    index_path.write_text(text, encoding="utf-8")


def main():
    repo_root = Path(__file__).resolve().parents[1]

    parser = argparse.ArgumentParser(description="Generate public-safe eCRF dashboard data from the local Excel workbook.")
    parser.add_argument("--input", default=str(repo_root / "private" / "query_dataset.xlsx"), help="Path to the local source workbook.")
    parser.add_argument("--output-js", default=str(repo_root / "ecrf-data.js"), help="Path to the generated ecrf-data.js file.")
    parser.add_argument("--output-json", default=str(repo_root / "ecrf-data.json"), help="Path to the generated ecrf-data.json file.")
    parser.add_argument("--index", default=str(repo_root / "ecrf" / "index.html"), help="Path to the dashboard index.html for cache-busting version updates.")
    parser.add_argument("--version", default=None, help="Optional explicit asset version token. Defaults to current UTC timestamp.")
    args = parser.parse_args()

    workbook_path = Path(args.input)
    output_js = Path(args.output_js)
    output_json = Path(args.output_json)
    index_path = Path(args.index)

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    payload = build_payload(workbook_path)
    output_js.parent.mkdir(parents=True, exist_ok=True)
    output_json.parent.mkdir(parents=True, exist_ok=True)

    write_outputs(payload, output_js, output_json)

    version_token = args.version or dt.datetime.now().strftime("%Y%m%d-%H%M%S")
    bump_asset_versions(index_path, version_token)

    meta = payload["META"]
    print("eCRF dashboard data refreshed successfully.")
    print(f"Workbook: {workbook_path}")
    print(f"Rows: {meta['totalQueries']}")
    print(f"Sites: {meta['activeSites']}")
    print(f"Participants: {meta['totalParticipants']}")
    print(f"Open: {meta['openQueries']}")
    print(f"Resolved: {meta['resolvedQueries']}")
    print(f"Not entered: {meta['notEnteredQueries']}")
    print(f"Last batch: {meta['lastQuerySent']}")
    print(f"Version token: {version_token}")


if __name__ == "__main__":
    main()
